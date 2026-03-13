import json
import re
from urllib.parse import quote
from collections import defaultdict
from io import BytesIO

import requests
import urllib3
import logging
import random
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill

from interface.utils import get_kibana_url, sample_tasks_with_dependencies

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

from django.core.cache import cache
from rest_framework.decorators import api_view
from rest_framework.response import Response
from django.http import JsonResponse, Http404
from django.shortcuts import get_object_or_404
from django.utils import timezone
from django.db.models import Q
from django.db import transaction
from datetime import timedelta, datetime
from django.urls import reverse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_GET, require_POST


from .models import Competition, LabLevel, Lab, LabTask, Answers, TeamCompetition2Team, TeamCompetition2TeamsAndUsers, User, TeamCompetition, LabTasksType, Kkz, Platoon, LabType, KkzPreview, Competition2User, TaskChecking
from .serializers import LabLevelSerializer, LabTaskSerializer
from .api_utils import get_issue
from .task_answer_parsing import parse_answer_choices
from .config import get_pnet_base_dir, get_pnet_url, get_web_url
from .utils import get_pnet_lab_name
from .eveFunctions import create_pnet_lab_session_common


def get_lab_type_code_from_display(lab_type_display):
    if not lab_type_display:
        return None
    
    for code, display in LabType.choices:
        if display == lab_type_display:
            return code
    return None


def find_lab_by_name_and_type(lab_name, lab_type_code=None):
    try:
        # Сначала пытаемся найти по slug
        if lab_type_code:
            return Lab.objects.get(slug=lab_name, lab_type=lab_type_code)
        else:
            return Lab.objects.get(slug=lab_name)
    except Lab.DoesNotExist:
        # Если не найдено по slug, пытаемся по name
        if lab_type_code:
            return Lab.objects.get(name=lab_name, lab_type=lab_type_code)
        else:
            return Lab.objects.get(name=lab_name)


@api_view(['GET'])
def get_time(request, instance_type, instance_id):  # pragma: no cover
    try:
        # Fetch the competition by ID
        if instance_type == 'kkz':
            instance = Kkz.objects.get(id=instance_id)
        else:  # competition
            instance = Competition.objects.get(id=instance_id)

        # Calculate remaining time
        remaining_time = instance.finish - timezone.now()
        if remaining_time < timedelta(0):
            remaining_time = timedelta(0)

        hours, remainder = divmod(int(remaining_time.total_seconds()), 3600)
        minutes, seconds = divmod(remainder, 60)

        return Response({
            'hours': hours,
            'minutes': minutes,
            'seconds': seconds,
        })

    except (Competition.DoesNotExist, Kkz.DoesNotExist):
        return Response({'error': f'{instance_type.upper()} not found'}, status=404)


def format_timedelta_ru(delta):
    if not delta:
        return ""
    
    days = delta.days
    remaining_seconds = delta.seconds
    hours, remainder = divmod(remaining_seconds, 3600)
    minutes, seconds = divmod(remainder, 60)
    
    parts = []
    
    if days > 0:
        if days == 1:
            day_word = "день"
        elif 2 <= days <= 4:
            day_word = "дня"
        else:
            day_word = "дней"
        parts.append(f"{days} {day_word}")
    
    # Форматируем время
    time_str = f"{hours:02d}:{minutes:02d}:{seconds:02d}"
    
    if parts:
        return f"{', '.join(parts)}, {time_str}"
    else:
        return time_str


def make_solution_data(solution, competition):
    user = User.objects.get(pk=solution["user_id"])
    return {
        "pos": 0,
        "user_first_name": user.first_name,
        "user_last_name": user.last_name,
        "user_platoon": str(user.platoon),
        "spent": format_timedelta_ru(solution["datetime"] - competition.start) if solution["datetime"] else "",
        "datetime": solution["datetime"].strftime("%H:%M:%S") if solution["datetime"] else "",
        "raw_datetime": solution["datetime"],
        "team_name": "",
        "user_id": solution["user_id"]
    }


def get_competition_by_slug(slug):
    try:
        return TeamCompetition.objects.get(slug=slug), True
    except TeamCompetition.DoesNotExist:
        return get_object_or_404(Competition, slug=slug), False


def get_competition_solutions_data(competition, is_team_competition=False):
    total_tasks = competition.num_tasks

    all_individual_users = User.objects.filter(
        platoon__in=competition.platoons.all()) | competition.non_platoon_users.all()

    if is_team_competition:
        team_user_ids = User.objects.filter(team__in=competition.teams.all()).values_list("id", flat=True)
        all_individual_users = all_individual_users.exclude(id__in=team_user_ids)

    individual_filter = Q(user__isnull=False) & (
            Q(user__platoon__in=competition.platoons.all()) |
            Q(user__in=competition.non_platoon_users.all())
    )
    individual_answers_qs = Answers.objects.filter(
        individual_filter,
        lab=competition.lab,
        datetime__lte=competition.finish,
        datetime__gte=competition.start
    )

    team_answers_qs = Answers.objects.none()
    if is_team_competition:
        team_answers_qs = Answers.objects.filter(
            team__in=competition.teams.all(),
            lab=competition.lab,
            datetime__lte=competition.finish,
            datetime__gte=competition.start,
            team__isnull=False
        )

    individual_data = {}
    for answer in individual_answers_qs:
        uid = answer.user.id
        if uid not in individual_data:
            individual_data[uid] = {
                'answers': [],
                'progress': 0,
                'raw_datetime': answer.datetime
            }
        individual_data[uid]['answers'].append(answer)
        if total_tasks > 0:
            individual_data[uid]['progress'] = len(individual_data[uid]['answers'])
        else:
            individual_data[uid]['progress'] = 1
        if answer.datetime > individual_data[uid]['raw_datetime']:
            individual_data[uid]['raw_datetime'] = answer.datetime

    for user in all_individual_users:
        if user.id not in individual_data:
            individual_data[user.id] = {
                'answers': [],
                'progress': 0,
                'raw_datetime': None
            }

    team_data = {}
    if is_team_competition:
        for answer in team_answers_qs:
            tid = answer.team.id
            if tid not in team_data:
                team_data[tid] = {
                    'answers': [],
                    'progress': 0,
                    'raw_datetime': answer.datetime,
                    'team': answer.team
                }
            team_data[tid]['answers'].append(answer)
            if total_tasks > 0:
                team_data[tid]['progress'] = len(team_data[tid]['answers'])
            else:
                team_data[tid]['progress'] = 1
            if answer.datetime > team_data[tid]['raw_datetime']:
                team_data[tid]['raw_datetime'] = answer.datetime

        for team in competition.teams.all():
            if team.id not in team_data:
                team_data[team.id] = {
                    'answers': [],
                    'progress': 0,
                    'raw_datetime': None,
                    'team': team
                }

    return {
        'competition': competition,
        'individual_data': individual_data,
        'team_data': team_data,
        'total_tasks': total_tasks,
        'all_individual_users': all_individual_users
    }


@api_view(['GET'])
def get_solutions(request, slug):
    competition, is_team_competition = get_competition_by_slug(slug)

    data = get_competition_solutions_data(competition, is_team_competition)
    individual_data = data['individual_data']
    team_data = data['team_data']
    total_tasks = data['total_tasks']

    solutions_data = []

    for uid, user_data in individual_data.items():
        comp2user = Competition2User.objects.filter(competition=competition, user_id=uid).first()
        if not comp2user or not comp2user.joined:
            continue
        dummy_solution = {
            "user_id": uid,
            "datetime": user_data['raw_datetime']
        }
        sol = make_solution_data(dummy_solution, competition)
        sol["progress"] = user_data['progress']

        if comp2user:
            max_tasks = comp2user.tasks.count()
        else:
            max_tasks = total_tasks or competition.tasks.count() or 1

        sol['max_tasks'] = max_tasks
        sol['total_tasks'] = max_tasks
        solutions_data.append(sol)

    if is_team_competition:
        for tid, t_data in team_data.items():
            team_obj = t_data['team']
            team_record = TeamCompetition2Team.objects.filter(competition=competition, team=team_obj).first()
            if not team_record or not team_record.joined:
                continue
            team_max_tasks = team_record.tasks.count() if team_record else (total_tasks or competition.tasks.count() or 1)

            for user in team_obj.users.all():
                dummy_solution = {
                    "user_id": user.id,
                    "datetime": t_data['raw_datetime']
                }
                sol = make_solution_data(dummy_solution, competition)
                sol["progress"] = t_data['progress']
                sol["team_name"] = team_obj.name
                sol["max_tasks"] = team_max_tasks
                sol["total_tasks"] = team_max_tasks
                solutions_data.append(sol)

    solutions_data.sort(key=lambda x: (-x["progress"], x["raw_datetime"] or timezone.now(), x["team_name"]))

    pos = 1
    for sol in solutions_data:
        sol["pos"] = pos
        pos += 1

    comp2users = Competition2User.objects.filter(competition=competition)
    max_assigned = 0
    if comp2users.exists():
        max_assigned = max((c.tasks.count() for c in comp2users), default=0)
    else:
        team_records = TeamCompetition2Team.objects.filter(competition=competition) if is_team_competition else TeamCompetition2Team.objects.none()
        if team_records.exists():
            max_assigned = max((t.tasks.count() for t in team_records), default=0)

    computed_total_tasks = total_tasks or max_assigned or competition.tasks.count() or 1

    response = {
        "solutions": solutions_data,
        "max_total_progress": competition.participants,
        "total_progress": sum(solution["progress"] for solution in solutions_data),
        "total_tasks": computed_total_tasks
    }
    if total_tasks:
        response["max_total_progress"] = competition.participants * total_tasks

    return JsonResponse(response)


def update_user_aggregated(user_aggregated, uid, user_data, competition, progress):
    if uid not in user_aggregated:
        try:
            user = User.objects.get(id=uid)
        except User.DoesNotExist:
            return

        user_aggregated[uid] = {
            'user_id': uid,
            'user': user,
            'progress': 0,
            'max_tasks': 0,
            'latest_datetime': None,
            'competition': competition
        }
    user_aggregated[uid]['progress'] += progress

    comp2user = Competition2User.objects.filter(
        competition=competition,
        user_id=uid
    ).first()

    if comp2user:
        user_aggregated[uid]['max_tasks'] += comp2user.tasks.count()

    raw_dt = user_data.get('raw_datetime') or user_data.get('datetime') or None
    if raw_dt and (not user_aggregated[uid]['latest_datetime'] or raw_dt > user_aggregated[uid]['latest_datetime']):
        user_aggregated[uid]['latest_datetime'] = raw_dt


@api_view(['GET'])
def get_kkz_solutions(request, kkz_id):
    try:
        kkz = Kkz.objects.get(id=kkz_id)
    except Kkz.DoesNotExist:
        return JsonResponse({'error': 'KKZ not found'}, status=404)

    competitions = Competition.objects.filter(kkz=kkz)

    if not competitions.exists():
        return JsonResponse({
            'solutions': [],
            'total_tasks': 0,
            'max_total_progress': 0,
            'total_progress': 0
        })

    user_aggregated = {}

    for competition in competitions:
        is_team_competition = hasattr(competition, 'teamcompetition')
        comp_data = get_competition_solutions_data(competition, is_team_competition)

        for uid, user_data in comp_data['individual_data'].items():
            update_user_aggregated(user_aggregated, uid, user_data, competition, user_data['progress'])

        if is_team_competition:
            for tid, t_data in comp_data['team_data'].items():
                team_obj = t_data['team']
                for user in team_obj.users.all():
                    update_user_aggregated(user_aggregated, user.id, t_data, competition, t_data['progress'])

    solutions_data = []
    total_max_tasks = 0

    for uid, data in user_aggregated.items():
        total_max_tasks += data['max_tasks']

        dummy_solution = {
            "user_id": uid,
            "datetime": data['latest_datetime']
        }
        sol = make_solution_data(dummy_solution, data['competition'])
        sol["progress"] = data['progress']
        sol["max_tasks"] = data['max_tasks']
        solutions_data.append(sol)

    solutions_data.sort(key=lambda x: (-x['progress'], x['raw_datetime'] if x['raw_datetime'] else timezone.now()))

    for pos, sol in enumerate(solutions_data, 1):
        sol['pos'] = pos

    return JsonResponse({
        'solutions': solutions_data,
        'max_total_progress': total_max_tasks,
        'total_progress': sum(s['progress'] for s in solutions_data)
    })


@api_view(['POST'])
def export_grades_xlsx(request):
    try:
        if hasattr(request, 'data'):
            data = request.data
        else:
            body = request.body.decode('utf-8')
            data = json.loads(body)
        
        grades_data = data.get('grades', [])
        instance_type = data.get('type')
        slug = data.get('slug')
        
        if not grades_data:
            return JsonResponse({'error': 'No grades data provided'}, status=400)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Оценки"
        headers = ['Фамилия', 'Имя', 'Оценка', 'Позиция', 'Решено заданий']
        ws.append(headers)
        
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        for grade_item in grades_data:
            row = [
                grade_item.get('last_name', ''),
                grade_item.get('first_name', ''),
                grade_item.get('grade', ''),
                grade_item.get('position', ''),
                grade_item.get('tasks_solved', 0)
            ]
            ws.append(row)
        
        column_widths = {
            'A': 20,  # Фамилия
            'B': 20,  # Имя
            'C': 10,  # Оценка
            'D': 10,  # Позиция
            'E': 15   # Решено заданий
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        for row_num in range(2, ws.max_row + 1):
            ws[f'A{row_num}'].alignment = Alignment(horizontal="center")
            ws[f'B{row_num}'].alignment = Alignment(horizontal="center")
            ws[f'C{row_num}'].alignment = Alignment(horizontal="center")
            ws[f'D{row_num}'].alignment = Alignment(horizontal="center")
            ws[f'E{row_num}'].alignment = Alignment(horizontal="center")
                
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        from django.http import HttpResponse
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        # Filename is generated on client side, so we just set a simple one
        response['Content-Disposition'] = 'attachment; filename="grades.xlsx"'
        
        return response
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        logging.error(f"Error exporting grades: {str(e)}")
        return JsonResponse({'error': f'Error exporting grades: {str(e)}'}, status=500)


@api_view(['POST'])
def save_grades(request):
    """Сохранить оценки в Competition2User / TeamCompetition2Team."""
    if not request.user.is_authenticated or not request.user.is_staff:
        return JsonResponse({'error': 'Forbidden'}, status=403)
    try:
        data = getattr(request, 'data', None) if isinstance(getattr(request, 'data', None), dict) else None
        if not data or 'grades' not in data:
            raw = request.body
            if raw:
                data = json.loads(raw.decode('utf-8') if isinstance(raw, bytes) else raw)
        if not data:
            return JsonResponse({'error': 'No data'}, status=400)
        slug = data.get('slug')
        instance_type = data.get('type')
        grades_data = data.get('grades') or []

        if not slug or instance_type != 'competition':
            return JsonResponse({'error': 'slug and type=competition required'}, status=400)
        competition, is_team_competition = get_competition_by_slug(slug)
        updated = 0
        # Участники соревнования для fallback по ФИ
        comp_user_ids = set(
            Competition2User.objects.filter(competition=competition).values_list('user_id', flat=True)
        )
        comp_team_user_ids = set()
        if is_team_competition:
            comp_team_user_ids = set(
                User.objects.filter(team__in=competition.teams.all()).values_list('id', flat=True)
            )

        for item in grades_data:
            grade_val = item.get('grade')
            if grade_val is None:
                continue
            try:
                grade = int(grade_val)
            except (TypeError, ValueError):
                continue
            if grade not in (2, 3, 4, 5):
                continue

            user_id = None
            raw_uid = item.get('user_id')
            if raw_uid not in (None, ''):
                try:
                    user_id = int(raw_uid)
                except (TypeError, ValueError):
                    pass
            if user_id is None:
                last_name = (item.get('user_last_name') or item.get('last_name') or '').strip()
                first_name = (item.get('user_first_name') or item.get('first_name') or '').strip()
                if last_name or first_name:
                    q = User.objects.filter(last_name__iexact=last_name, first_name__iexact=first_name)
                    allowed_ids = comp_user_ids | comp_team_user_ids
                    if allowed_ids:
                        q = q.filter(id__in=allowed_ids)
                    u = q.first()
                    if u:
                        user_id = u.id
            if user_id is None:
                continue

            if is_team_competition:
                team_record = TeamCompetition2Team.objects.filter(
                    competition=competition, team__users=user_id
                ).first()
                if team_record:
                    TeamCompetition2Team.objects.filter(pk=team_record.pk).update(grade=grade)
                    updated += 1
            else:
                n = Competition2User.objects.filter(
                    competition=competition, user_id=user_id
                ).update(grade=grade)
                if n > 0:
                    updated += 1
                else:
                    comp2user, _ = Competition2User.objects.get_or_create(
                        competition=competition, user_id=user_id, defaults={'joined': True}
                    )
                    Competition2User.objects.filter(pk=comp2user.pk).update(grade=grade)
                    updated += 1
        return JsonResponse({'ok': True, 'updated': updated})
    except json.JSONDecodeError as e:
        return JsonResponse({'error': 'Invalid JSON: ' + str(e)}, status=400)
    except Exception as e:
        logging.exception("Error saving grades")
        return JsonResponse({'error': str(e)}, status=500)


@api_view(['GET'])
def my_grade(request, slug):
    """Вернуть оценку текущего пользователя за лабу."""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'Unauthorized'}, status=401)
    competition, is_team = get_competition_by_slug(slug)
    if is_team:
        grade = TeamCompetition2Team.objects.filter(
            competition=competition, team__users=request.user
        ).values_list('grade', flat=True).first()
    else:
        grade = Competition2User.objects.filter(
            competition=competition, user=request.user
        ).values_list('grade', flat=True).first()
    response = JsonResponse({'grade': grade})
    response['Cache-Control'] = 'no-store, no-cache, must-revalidate'
    response['Pragma'] = 'no-cache'
    return response


@api_view(['GET'])
def load_levels(request, lab_name):  # pragma: no cover
    try:
        lab_type_display = request.GET.get('lab_type_display')
        lab_type_code = get_lab_type_code_from_display(lab_type_display)
        
        lab = find_lab_by_name_and_type(lab_name, lab_type_code)
        levels = LabLevel.objects.filter(lab=lab)
        serializer = LabLevelSerializer(levels, many=True)
        
        return Response(serializer.data)
    except Lab.DoesNotExist:
        return Response({"error": "Lab not found"}, status=404)


@api_view(['GET'])
def load_tasks(request, lab_name):  # pragma: no cover
    try:
        lab_type_display = request.GET.get('lab_type_display')
        lab_type_code = get_lab_type_code_from_display(lab_type_display)
        
        lab = find_lab_by_name_and_type(lab_name, lab_type_code)
        tasks = LabTask.objects.filter(lab=lab)
        serializer = LabTaskSerializer(tasks, many=True)
        
        return Response(serializer.data)
    except Lab.DoesNotExist:
        return Response({"error": "Lab not found"}, status=404)


def change_iso_timezone(utc_time):  # pragma: no cover
    print(utc_time)
    if utc_time[-1] == 'Z':
        utc_time = utc_time[:-1]
    return datetime.fromisoformat(utc_time) + timedelta(hours=3)


def update_instance_time(instance, action, minutes=15):
    if action == "start":
        time = timezone.now()
        field = "start"
        message = "started"
    elif action == "end":
        time = timezone.now()
        field = "finish"
        message = "ended"
    elif action == "resume":
        current_finish = getattr(instance, 'finish', None)
        now = timezone.now()
        if current_finish > now:
            time = current_finish + timedelta(minutes=minutes)
        else:
            time = now + timedelta(minutes=minutes)
        field = "finish"
        message = f"resumed for {minutes} minutes"
    else:
        return None, "Unknown action"

    setattr(instance, field, time)
    instance.save()

    if isinstance(instance, Kkz):
        Competition.objects.filter(kkz=instance).update(**{field: time})

    cache.set("competitions_update", True, timeout=60)
    return message, None


def delete_competition_from_platform(competition):
    """
    Удаляет соревнование с платформы: удаляет competition2user, competition2team,
    сессии сегментов (TeamCompetition2TeamsAndUsers — возврат воркспейсов участникам),
    затем помечает competition как deleted.
    """
    from .pnet_session_manager import with_pnet_session_if_needed

    competitions2users = Competition2User.objects.filter(competition=competition, deleted=False)
    competitions2teams = TeamCompetition2Team.objects.filter(competition=competition, deleted=False)

    # Сессии сегментов (лаба в воркспейсе мастера): удалить лабу и вернуть участников в свои воркспейсы
    try:
        team_competition = TeamCompetition.objects.get(pk=competition.pk)
        segment_sessions = TeamCompetition2TeamsAndUsers.objects.filter(
            team_competition=team_competition, deleted=False
        )
    except TeamCompetition.DoesNotExist:
        segment_sessions = []

    def _delete_segment_sessions_operation():
        for session in segment_sessions:
            session.delete_from_platform()

    if segment_sessions:
        with_pnet_session_if_needed(competition.lab, _delete_segment_sessions_operation)

    # Удаляем competition2user
    if competitions2users.exists():
        def _delete_competitions2users_operation():
            for competition2user in competitions2users:
                competition2user.delete_from_platform()
        
        with_pnet_session_if_needed(competition.lab, _delete_competitions2users_operation)

    # Удаляем competition2team
    if competitions2teams.exists():
        def _delete_competition2team_operation():
            for competition2team in competitions2teams:
                competition2team.delete_from_platform()
        
        with_pnet_session_if_needed(competition.lab, _delete_competition2team_operation)
    
    # Помечаем competition как deleted
    competition.deleted = True
    competition.save()
    
    return "deleted from platform"


def delete_kkz_from_platform(kkz):
    """
    Удаляет ККЗ с платформы: удаляет все экзамены с платформы, если они завершены и платформа PN или CMD
    """
    now = timezone.now()
    deleted_count = 0
    
    competitions = Competition.objects.filter(
        kkz=kkz,
        finish__lte=now,
        deleted=False,
        lab__platform__in=["PN", "CMD"]
    ).select_related('lab')
    
    for competition in competitions:
        delete_competition_from_platform(competition)
        deleted_count += 1
    
    return f"deleted {deleted_count} competition(s) from platform"


@api_view(['POST'])
def press_button(request, action):
    try:
        minutes = request.data.get('minutes', 15)
        print(f"Action: {action}, Minutes: {minutes}")
        print(f"Request data: {request.data}")

        slug = request.data.get('slug')
        kkz_id = request.data.get('kkz_id')

        if kkz_id:
            instance = get_object_or_404(Kkz, id=kkz_id)
            redirect_url = reverse('interface:kkz-detail', kwargs={'pk': kkz_id})
        elif slug:
            instance = get_object_or_404(Competition, slug=slug)
            redirect_url = reverse('interface:competition-detail', kwargs={'slug': slug})
        else:
            return JsonResponse({"error": "Missing slug or kkz_id"}, status=400)

        # Обработка действия удаления с платформы
        if action == "delete":
            if isinstance(instance, Kkz):
                message = delete_kkz_from_platform(instance)
                instance_type = "KKZ"
            else:
                message = delete_competition_from_platform(instance)
                instance_type = "Competition"
            
            cache.set("competitions_update", True, timeout=60)
            return JsonResponse({
                "message": f"{instance_type} {message}",
                "redirect_url": redirect_url
            }, status=200)

        message, error = update_instance_time(instance, action, minutes)

        if error:
            return JsonResponse({"error": error}, status=400)

        instance_type = "KKZ" if isinstance(instance, Kkz) else "Competition"
        return JsonResponse({
            "message": f"{instance_type} {message}",
            "redirect_url": redirect_url
        }, status=200)

    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f"Server error in press_button: {e}")
        return JsonResponse({"error": str(e)}, status=500)


@api_view(['GET'])
def check_updates(request):  # pragma: no cover
    last_update = cache.get("competitions_update", False)
    if last_update:
        cache.delete("competitions_update")
    return JsonResponse({"update_required": last_update})


@api_view(['GET'])
def check_availability(request, slug):  # pragma: no cover
    try:
        competition = Competition.objects.get(slug=slug)
        available = competition.finish > timezone.now()
        return JsonResponse({"available": available})
    except Competition.DoesNotExist:
        return JsonResponse({"error": "Competition not found"}, status=404)


@api_view(['GET'])
def get_users_in_platoons(request):  # pragma: no cover
    platoon_ids = request.GET.get('platoons', '')

    if platoon_ids:
        ids = [int(x) for x in platoon_ids.split(',') if x.isdigit()]
        users = User.objects.filter(platoon__number__in=ids).order_by('username')
        print(users)
        data = [{"id": user.pk, "login": user.username} for user in users]
    else:
        data = []
    return JsonResponse(data, safe=False)


def create_var_text(second_name):
    new_var = rf"""/testdir/ 1 1 1 1 1 1 1 1 1 1
./ {second_name}d1 drwxrwxrwxm-- admin admin Секретно:Низкий:Нет:0x0
{second_name}d1/ {second_name}d2 drwxrwx---m-- admin admin Секретно:Низкий:Нет:0x0
{second_name}d1/ {second_name}f1 -rwx------m-- admin admin Секретно:Низкий:Нет:0x0
{second_name}d1/ {second_name}f3 -rwx------m-- admin admin Секретно:Низкий:Нет:0x0"""
    return new_var


def parse_request_data(request):
    try:
        return json.loads(request.body.decode('utf-8'))
    except (ValueError, TypeError):
        raise


def get_issue_tasks(issue):
    if issue.competition.lab.tasks_type == LabTasksType.CLASSIC:
        tasks = [task.task_id for task in issue.tasks.all()]
    else:
        tasks = [{'id': task.task_id, **task.json_config} for task in issue.tasks.all()]
    return tasks


@api_view(['GET'])
def start_lab(request):
    if request.method == 'GET':
        issue, error_response = get_issue(
            parse_request_data(request),
            {'competition__finish__gt': timezone.now()}
        )
        logging.info(f"payload: {parse_request_data(request)}")
        if error_response:
            return error_response

        response_data = {
            "tasks": get_issue_tasks(issue)
        }
        if hasattr(issue, 'user'):
            response_data["task"] = create_var_text(issue.user.last_name),
        if hasattr(issue, 'level') and issue.level:
            response_data["variant"] = issue.level.level_number
        if issue.competition.lab.answer_flag:
            response_data["flag"] = issue.competition.lab.answer_flag
        if hasattr(issue, 'team') and issue.team:
            response_data["team"] = [user.pnet_login for user in issue.team.users.all()]
        response_data["type"] = issue.competition.lab.lab_type
        return JsonResponse(response_data)


@api_view(['POST'])
def end_lab(request):
    if request.method == 'POST':
        issue, error_response = get_issue(
            parse_request_data(request),
            {'competition__start__lt': timezone.now()}
        )
        if error_response:
            return error_response

        if issue.tasks.count() > 0:
            return JsonResponse({'message': 'Task finished'})

        if hasattr(issue, 'user'):
            Answers.objects.update_or_create(
                lab=issue.competition.lab, 
                user=issue.user,
                lab_task=None,
                defaults={'datetime': timezone.now()}
            )
        elif hasattr(issue, 'team'):
            Answers.objects.update_or_create(
                lab=issue.competition.lab,
                team=issue.team,
                lab_task=None,
                defaults={'datetime': timezone.now()}
            )

        return JsonResponse({'message': 'Task finished'})


@csrf_exempt
@api_view(['POST'])
def get_pnet_auth(request):
    """Аутентифицирует пользователя в PNET и возвращает cookies"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'User not authenticated'}, status=401)

    user = request.user
    if not user.pnet_login or not user.pnet_password:
        return JsonResponse({'error': 'User PNET credentials not configured'}, status=400)

    try:
        # Тот же URL, что и для create_pnet_lab_session_with_console
        pnet_url = get_pnet_url() or f"{get_web_url()}/pnetlab"
        session = requests.Session()
        session.verify = False

        # Получаем XSRF-TOKEN
        preflight_response = session.get(f"{pnet_url}/store/public/auth/login/login", timeout=10)
        if preflight_response.status_code not in [200, 202]:
            return JsonResponse({'error': 'Failed to connect to PNET'}, status=500)

        xsrf_token = session.cookies.get('XSRF-TOKEN', '')

        # Аутентификация
        headers = {'Content-Type': 'application/json;charset=UTF-8', 'X-Requested-With': 'XMLHttpRequest'}
        if xsrf_token:
            headers['X-XSRF-TOKEN'] = xsrf_token

        login_response = session.post(
            f"{pnet_url}/store/public/auth/login/login",
            headers=headers,
            json={'username': user.pnet_login, 'password': user.pnet_password, 'html': '1', 'captcha': ''},
            timeout=10
        )

        if login_response.status_code not in [200, 201, 202]:
            return JsonResponse({'error': 'PNET authentication failed'}, status=401)

        # Проверка JSON ответа для статуса 202
        if login_response.status_code == 202:
            try:
                if not login_response.json().get('result', False):
                    return JsonResponse({'error': 'PNET authentication failed'}, status=401)
            except (ValueError, KeyError):
                pass  # Продолжаем, возможно cookies установлены

        # Возвращаем cookies и устанавливаем их в ответе
        cookies_dict = {cookie.name: cookie.value for cookie in session.cookies}
        cookie_names = list(cookies_dict.keys())
        logger = logging.getLogger(__name__)
        logger.info(
            "get_pnet_auth: success for user=%s pnet_login=%s setting %s cookies: %s",
            user.username,
            getattr(user, "pnet_login", None),
            len(cookie_names),
            cookie_names,
        )
        response = JsonResponse({'success': True, 'cookies': cookies_dict, 'xsrf_token': xsrf_token})

        for cookie in session.cookies:
            logger.info(f"Setting cookie {cookie.name}={cookie.value} domain={cookie.domain} path={cookie.path}")
            response.set_cookie(
                cookie.name,
                cookie.value,
                path='/',  # Принудительно устанавливаем корневой путь
                secure=False,  # Отключаем secure для HTTP
                httponly=False,  # Разрешаем JavaScript доступ
                samesite='Lax'
            )

        return response

    except requests.exceptions.Timeout:
        return JsonResponse({'error': 'PNET request timeout'}, status=500)
    except requests.exceptions.ConnectionError:
        return JsonResponse({'error': 'Failed to connect to PNET'}, status=500)
    except Exception as e:
        return JsonResponse({'error': f'Authentication error: {str(e)}'}, status=500)


@api_view(['GET'])
def check_kibana_auth_status(request):
    """Проверяет статус аутентификации пользователя в Kibana"""
    if not request.user.is_authenticated:
        return JsonResponse({'authenticated': False, 'error': 'User not authenticated'}, status=401)

    user = request.user
    if not user.pnet_login or not user.pnet_password:
        return JsonResponse({'authenticated': False, 'error': 'User credentials not configured'}, status=400)

    try:
        kibana_url = get_kibana_url()
        session = requests.Session()
        session.verify = False

        # Проверяем статус аутентификации через API Kibana
        headers = {
            'Accept': 'application/json',
            'kbn-version': '9.1.2',
            'x-elastic-internal-origin': 'Kibana'
        }

        # Пробуем получить информацию о текущем пользователе
        status_response = session.get(
            f"{kibana_url}/internal/security/me",
            headers=headers,
            timeout=5
        )

        # Если получили 200 - пользователь аутентифицирован
        if status_response.status_code == 200:
            try:
                user_info = status_response.json()
                return JsonResponse({
                    'authenticated': True,
                    'username': user_info.get('username', user.pnet_login)
                })
            except (ValueError, KeyError):
                pass

        # Если получили 401/403 - пользователь не аутентифицирован
        return JsonResponse({'authenticated': False})

    except requests.exceptions.Timeout:
        return JsonResponse({'authenticated': False, 'error': 'Kibana request timeout'}, status=500)
    except requests.exceptions.ConnectionError:
        return JsonResponse({'authenticated': False, 'error': 'Failed to connect to Kibana'}, status=500)
    except Exception as e:
        return JsonResponse({'authenticated': False, 'error': f'Check error: {str(e)}'}, status=500)


@csrf_exempt
@api_view(['POST'])
def get_kibana_auth(request):
    """Аутентифицирует пользователя в Kibana и возвращает cookies"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'User not authenticated'}, status=401)

    user = request.user
    if not user.pnet_login or not user.pnet_password:
        return JsonResponse({'error': 'User credentials not configured'}, status=400)

    try:
        kibana_url = get_kibana_url()
        session = requests.Session()
        session.verify = False

        # Получаем главную страницу Kibana для получения версии
        main_page_response = session.get(f"{kibana_url}/", timeout=10)
        if main_page_response.status_code not in [200, 401, 403]:
            return JsonResponse({'error': 'Failed to connect to Kibana'}, status=500)

        # Извлекаем версию Kibana из заголовков
        kbn_version = main_page_response.headers.get('kbn-version', '9.1.2')

        # Аутентификация через внутренний API Kibana
        headers = {
            'Content-Type': 'application/json',
            'Accept': '*/*',
            'Accept-Encoding': 'gzip, deflate, br, zstd',
            'Accept-Language': 'ru,en;q=0.9',
            'Connection': 'keep-alive',
            'Origin': kibana_url,
            'Referer': f"{kibana_url}/login?msg=LOGGED_OUT",
            'Sec-Fetch-Dest': 'empty',
            'Sec-Fetch-Mode': 'cors',
            'Sec-Fetch-Site': 'same-origin',
            'User-Agent': 'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/138.0.0.0 YaBrowser/25.8.0.0 Safari/537.36',
            'kbn-build-number': '88427',
            'kbn-version': kbn_version,
            'x-elastic-internal-origin': 'Kibana'
        }

        login_data = {
            'providerType': 'basic',
            'providerName': 'basic',
            'currentURL': f"{kibana_url}/login?msg=LOGGED_OUT",
            'params': {
                'username': user.pnet_login,
                'password': user.pnet_password
            }
        }

        login_response = session.post(
            f"{kibana_url}/internal/security/login",
            headers=headers,
            json=login_data,
            timeout=10
        )

        if login_response.status_code not in [200, 201, 202]:
            return JsonResponse({'error': 'Kibana authentication failed'}, status=401)

        # Проверяем успешность аутентификации
        try:
            login_result = login_response.json()
            if not login_result.get('success', True):
                return JsonResponse({'error': 'Kibana authentication failed'}, status=401)
        except (ValueError, KeyError):
            pass  # Продолжаем, возможно cookies установлены

        # Возвращаем cookies и устанавливаем их в ответе
        cookies_dict = {cookie.name: cookie.value for cookie in session.cookies}
        response = JsonResponse({'success': True, 'cookies': cookies_dict})

        # Устанавливаем cookies в HTTP-ответе для автоматической синхронизации
        for cookie in session.cookies:
            response.set_cookie(
                cookie.name,
                cookie.value,
                path='/',
                secure=False,
                httponly=False,
                samesite='Lax'
            )

        return response

    except requests.exceptions.Timeout:
        return JsonResponse({'error': 'Kibana request timeout'}, status=500)
    except requests.exceptions.ConnectionError:
        return JsonResponse({'error': 'Failed to connect to Kibana'}, status=500)
    except Exception as e:
        return JsonResponse({'error': f'Kibana authentication error: {str(e)}'}, status=500)


def get_lab_path(competition, user, session_issue=None):
    """
    Возвращает путь к лабе для пользователя.

    Для сессии по сегментам (session_issue задан): лаба в воркспейсе master_user.
    Для командных участников: /base_path/{team.slug}/lab.unl
    Для одиночных участников: /base_path/{user.pnet_login}/lab.unl
    """
    base_path = get_pnet_base_dir().strip('/')
    lab_file_name = f"{get_pnet_lab_name(competition)}.unl"

    if session_issue is not None and getattr(session_issue, 'master_user', None):
        user_dir = session_issue.master_user.pnet_login
    else:
        # Определяем директорию пользователя
        user_dir = user.pnet_login  # По умолчанию - индивидуальная директория
        if isinstance(competition, TeamCompetition):
            # Для командных соревнований пытаемся получить slug команды
            try:
                team = get_user_team(competition, user)
                user_dir = team.slug
            except TeamCompetition2Team.DoesNotExist:
                pass  # Используем индивидуальную директорию

    return f"/{base_path}/{user_dir}/{lab_file_name}"


def _get_team_issue_for_user(competition, user):
    return TeamCompetition2Team.objects.select_related('team').filter(
        competition=competition,
        team__users=user
    ).first()


def _get_latest_team_issue(team_issue):
    return TeamCompetition2Team.objects.select_related('master_session_user').get(pk=team_issue.pk)


def _copy_request_cookies_to_session(session, request):
    for cookie_name, cookie_value in request.COOKIES.items():
        session.cookies.set(cookie_name, cookie_value)


def _get_master_session_id(pnet_url, cookie, xsrf, lab_path):
    """
    Получает session_id мастера через filter_session.
    Использует cookie текущего пользователя, чтобы не логиниться от имени мастера.
    
    Args:
        pnet_url: URL PNET сервера
        cookie: Cookie текущего пользователя
        xsrf: XSRF токен текущего пользователя
        lab_path: Путь к лабе
    
    Returns:
        tuple: (session_id, error_message)
    """
    from .eveFunctions import get_session_id_by_filter

    return get_session_id_by_filter(pnet_url, cookie, xsrf, lab_path)


def _create_master_session(pnet_url, master_user, lab_path, cookie, xsrf):
    """
    Создает мастер-сессию от имени master_user.
    
    Args:
        pnet_url: URL PNET сервера
        master_user: Пользователь-мастер команды
        lab_path: Путь к лабе
        cookie: Cookie текущего пользователя
        xsrf: XSRF токен текущего пользователя
    
    Returns:
        tuple: (session_id, error_message)
    """
    if not master_user or not master_user.pnet_login or not master_user.pnet_password:
        return None, "Master user PNET credentials not configured"

    from .eveFunctions import login_user_to_pnet, create_pnet_lab_session_common, get_session_id_by_filter

    session, xsrf_master = login_user_to_pnet(pnet_url, master_user.pnet_login, master_user.pnet_password)
    if not session:
        return None, "Failed to login as master user"

    success, message, _ = create_pnet_lab_session_common(
        pnet_url, master_user.pnet_login, lab_path, session.cookies, xsrf=xsrf_master
    )
    if not success:
        return None, message

    # Используем cookie текущего пользователя для получения session_id через filter
    return get_session_id_by_filter(pnet_url, cookie, xsrf, lab_path)


def _join_master_session(pnet_url, cookies, xsrf, lab_path, master_session_id):
    from .eveFunctions import join_session

    join_response = join_session(pnet_url, master_session_id, cookies, xsrf=xsrf)
    if join_response.status_code not in range(200, 400):
        err_preview = (join_response.text or '')[:300]
        logging.warning(
            'PNET join_session failed: status=%s body=%s',
            join_response.status_code,
            err_preview,
        )
        return False, f"Failed to join master session: {join_response.text}"

    return True, None


def _ensure_team_session(competition, user, pnet_url, cookies, xsrf):
    team_issue = _get_team_issue_for_user(competition, user)
    if not team_issue:
        return False, 'Team for user not found in competition', None

    latest_issue = _get_latest_team_issue(team_issue)
    master_user = latest_issue.master_session_user
    if not master_user:
        return False, 'Master user not configured for team', None

    lab_path = get_lab_path(competition, user)

    # Если текущий пользователь — мастер, не логиним его заново (инвалидировали бы его же сессию).
    # Создаём сессию лабы его же cookies из запроса.
    if user == master_user:
        from .eveFunctions import create_pnet_lab_session_common
        success, message, _ = create_pnet_lab_session_common(
            pnet_url, user.pnet_login, lab_path, cookies, xsrf=xsrf
        )
        if not success:
            return False, message, None
        TeamCompetition2Team.objects.filter(pk=team_issue.pk).update(master_session_created=True)
        return True, None, lab_path

    if not latest_issue.master_session_created:
        master_session_id, error_message = _create_master_session(pnet_url, master_user, lab_path, cookies, xsrf)
        if not master_session_id:
            return False, error_message, None
        TeamCompetition2Team.objects.filter(pk=team_issue.pk).update(master_session_created=True)
    else:
        master_session_id, error_message = _get_master_session_id(pnet_url, cookies, xsrf, lab_path)
        if not master_session_id:
            return False, error_message, None
        
    success, message = _join_master_session(pnet_url, cookies, xsrf, lab_path, master_session_id=master_session_id)
    if not success:
        return False, message, None

    return True, None, lab_path


def _ensure_segment_session(competition, user, session_issue, pnet_url, cookies, xsrf):
    """
    Лабу создаёт тот, кто первый открыл. Сохраняем session_id и владельца.
    Все остальные (включая мастера) джойнятся к этой сессии без 412 у мастера.
    """
    from .eveFunctions import create_pnet_lab_session_common

    master_user = session_issue.master_user
    if not master_user:
        return False, 'Master user not configured for session', None

    lab_path = get_lab_path(competition, master_user, session_issue=session_issue)
    session_issue.refresh_from_db()
    stored_session_id = session_issue.segment_pnet_session_id
    owner_id = session_issue.segment_pnet_session_owner_id

    if not stored_session_id:
        # Первый открывший: создаём лабу в его сессии
        if not user.pnet_login:
            return False, 'PNET login not configured for user', None
        create_ok, create_msg = create_pnet_lab_session_common(
            pnet_url, user.pnet_login, lab_path, cookies, xsrf
        )
        if not create_ok:
            return False, create_msg or 'Failed to create lab session', None
        master_session_id, err = _get_master_session_id(pnet_url, cookies, xsrf, lab_path)
        if not master_session_id:
            return False, err or 'Session not found after create', None
        session_issue.segment_pnet_session_id = str(master_session_id)
        session_issue.segment_pnet_session_owner_id = user.id
        session_issue.master_session_created = True
        session_issue.save(update_fields=['segment_pnet_session_id', 'segment_pnet_session_owner_id', 'master_session_created'])
        logging.info('PNET segment: user_id=%s created lab, session_id=%s', user.id, master_session_id)
        return True, None, lab_path

    if owner_id is not None and user.id == owner_id:
        logging.info('PNET segment: user_id=%s is session owner, skip join', user.id)
        return True, None, lab_path

    logging.info('PNET segment: user_id=%s joining session_id=%s', user.id, stored_session_id)
    success, message = _join_master_session(pnet_url, cookies, xsrf, lab_path, master_session_id=stored_session_id)
    if not success:
        logging.warning('PNET segment: user_id=%s join failed: %s', user.id, message)
        return False, message, None
    logging.info('PNET segment: user_id=%s join ok', user.id)
    return True, None, lab_path


@csrf_exempt
@api_view(['POST'])
def create_pnet_lab_session(request):
    """Создает сессию лабы в PNET после аутентификации"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'User not authenticated'}, status=401)

    slug = request.data.get('slug')
    if not slug:
        return JsonResponse({'error': 'Competition slug required'}, status=400)

    logger = logging.getLogger(__name__)
    try:
        from .api_utils import get_issue_for_user, create_lab_session_for_issue
        
        competition, _ = get_competition_by_slug(slug)
        user = request.user

        if not user.pnet_login:
            return JsonResponse({'error': 'User PNET login not configured'}, status=400)

        # Получаем issue для пользователя
        issue, error_response = get_issue_for_user(competition, user)
        if error_response:
            return error_response

        # Тот же URL, что и в get_pnet_auth, иначе cookies от другого хоста не подойдут
        pnet_url = get_pnet_url() or f"{get_web_url()}/pnetlab"
        session = requests.Session()
        session.verify = False

        # Копируем cookies из запроса (должны быть установлены после get_pnet_auth)
        request_cookie_names = list(request.COOKIES.keys()) if request.COOKIES else []
        logger.info(
            "create_pnet_lab_session: user=%s request.COOKIES names=%s",
            request.user.username,
            request_cookie_names,
        )
        _copy_request_cookies_to_session(session, request)
        session_cookie_names = list(session.cookies.keys())
        logger.info(
            "create_pnet_lab_session: after copy session.cookies names=%s",
            session_cookie_names,
        )
        xsrf_token = session.cookies.get('XSRF-TOKEN', '')

        # Создаем сессию лабы
        success, message, lab_path, pnet_code = create_lab_session_for_issue(
            competition, user, issue, pnet_url, session.cookies, xsrf_token
        )

        if not success:
            logger.warning(
                "create_pnet_lab_session: failed user=%s success=False message=%s pnet_code=%s cookies_sent=%s",
                request.user.username,
                message[:200] if message else None,
                pnet_code,
                list(session.cookies.keys()),
            )
            if '412' in str(message) or 'unauthorized' in str(message).lower() or 'session timed out' in str(message).lower():
                logging.warning('create_pnet_lab_session: join failed for user %s: %s', request.user.id, message)
                return JsonResponse({
                    'error': 'Сессия PNET не готова или истекла. Обновите страницу и откройте лабу снова.'
                }, status=503)
            if message == 'Team for user not found in competition':
                status_code = 404
            elif isinstance(pnet_code, int) and 400 <= pnet_code <= 599:
                status_code = pnet_code
            else:
                status_code = 500
            return JsonResponse({'error': message}, status=status_code)

        # Запускаем все ноды в топологии
        import time
        from .eveFunctions import get_lab_topology, turn_on_node
        from .lab_topology import LabTopology
        logger = logging.getLogger(__name__)
        topology_data = None
        for attempt in range(4):
            if attempt > 0:
                time.sleep(4)
            topology_data = get_lab_topology(pnet_url, session.cookies, xsrf_token)
            if topology_data and topology_data.get('code') == 200:
                break
        if not topology_data or topology_data.get('code') != 200:
            logging.warning('create_pnet_lab_session: topology failed for user %s after retries', request.user.id)
            return JsonResponse({
                'error': 'Не удалось загрузить топологию (сессия PNET). Обновите страницу и откройте лабу снова.'
            }, status=503)
        topology = LabTopology(topology_data)
        all_node_ids = topology.get_all_node_ids()
        if all_node_ids:
            logger.info(f'Starting {len(all_node_ids)} nodes for lab session: {all_node_ids}')
            for node_id in all_node_ids:
                ok, msg = turn_on_node(pnet_url, node_id, session.cookies, xsrf_token)
                if not ok:
                    logger.warning(f'Node {node_id} start failed: {msg}')
                else:
                    logger.debug(f'Node {node_id} started')

        path_q = quote(lab_path or '', safe='')
        redirect_url = f'/legacy/topology?path={path_q}' if path_q else '/legacy/topology'
        return JsonResponse({
            'success': True,
            'lab_path': lab_path,
            'redirect_url': redirect_url,
        })

    except Competition.DoesNotExist:
        return JsonResponse({'error': 'Competition not found'}, status=404)
    except requests.exceptions.Timeout:
        return JsonResponse({'error': 'PNET request timeout'}, status=500)
    except requests.exceptions.ConnectionError:
        return JsonResponse({'error': 'Failed to connect to PNET'}, status=500)
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.exception(f'Session creation error for user {request.user.username}: {str(e)}')
        return JsonResponse({'error': f'Session creation error: {str(e)}'}, status=500)


def get_user_team(team_competition, user):
    return team_competition.competition_teams.get(team__users=user).team


@csrf_exempt
@api_view(['POST'])
def create_pnet_lab_session_with_console(request):
    """Создает сессию лабы в PNET и возвращает ссылку на консоль SSH ноды"""
    slug = request.data.get('slug')
    if not slug:
        return JsonResponse({'error': 'Competition slug required'}, status=400)
    
    # Получаем node_name из request.data, если не указан - используем PnetSSHNodeName
    node_name = request.data.get('node_name')
    
    # Получаем username из request.data, если не передан - используем request.user
    username = request.data.get('username')
    if not username:
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Username required or user must be authenticated'}, status=400)
        username = request.user.username

    try:
        logger = logging.getLogger(__name__)
        logger.info(f'Creating CMD console session for slug: {slug}, username: {username}')
        
        from .api_utils import get_user_by_username, get_issue_for_user, create_lab_session_for_issue
        
        competition, _ = get_competition_by_slug(slug)
        
        # Получаем пользователя по username
        user = get_user_by_username(username)
        if not user:
            return JsonResponse({'error': f'User with username "{username}" not found'}, status=404)
        
        # Если node_name не передан в запросе, используем PnetSSHNodeName из лабы
        if not node_name:
            node_name = competition.lab.PnetSSHNodeName
        
        logger.info(f'Competition found: {competition}, Lab: {competition.lab}, SSH Node: {node_name}')

        if not user.pnet_login or not user.pnet_password:
            logger.error(f'User PNET credentials not configured for user: {user}')
            return JsonResponse({'error': 'User PNET credentials not configured'}, status=400)

        if not node_name:
            logger = logging.getLogger(__name__)
            logger.error(f'SSH node name not configured for lab: {competition.lab}')
            return JsonResponse({'error': 'SSH node name not configured for this lab'}, status=400)

        # Импортируем функции из eveFunctions
        from .eveFunctions import pf_login, get_lab_topology, get_guacamole_url, turn_on_node
        from .lab_topology import LabTopology

        # Получаем URL PNET
        pnet_url = get_pnet_url()
        
        # Получаем issue для пользователя
        issue, error_response = get_issue_for_user(competition, user)
        if error_response:
            return error_response
        
        # Логинимся в PNET
        cookies, xsrf_token = pf_login(pnet_url, user.pnet_login, user.pnet_password)

        # Создаём/джойнимся к сессии лабы
        success, message, lab_path, pnet_code= create_lab_session_for_issue(
            competition, user, issue, pnet_url, cookies, xsrf_token
        )
        
        if not success:
            if message == 'Team for user not found in competition':
                status_code = 404
            elif isinstance(pnet_code, int) and 400 <= pnet_code <= 599:
                status_code = pnet_code
            else:
                status_code = 500
            return JsonResponse({'error': message}, status=status_code)

        # Для сегментных сессий топологию и ноды управляем от имени мастер-пользователя
        if isinstance(issue, TeamCompetition2TeamsAndUsers) and issue.master_user:
            master_user = issue.master_user
            logger.info(f'Segment session: switching to master user {master_user.pnet_login} for topology')
            master_cookies, master_xsrf = pf_login(pnet_url, master_user.pnet_login, master_user.pnet_password)
            # Открываем сессию от имени мастера (join если уже есть, create если нет)
            issue.refresh_from_db()
            if issue.segment_pnet_session_id:
                from .eveFunctions import join_session
                join_session(pnet_url, issue.segment_pnet_session_id, master_cookies, master_xsrf)
            else:
                create_pnet_lab_session_common(pnet_url, master_user.pnet_login, lab_path, master_cookies, master_xsrf)
            cookies, xsrf_token = master_cookies, master_xsrf

        # Сессия создана; топологию запрашиваем с повторами
        import time
        topology_data = None
        for attempt in range(4):
            if attempt > 0:
                time.sleep(4)
            topology_data = get_lab_topology(pnet_url, cookies, xsrf_token)
            if topology_data and topology_data.get('code') == 200:
                break
        if not topology_data or topology_data.get('code') != 200:
            return JsonResponse({'error': 'Failed to get lab topology'}, status=500)

        topology = LabTopology(topology_data)
        logger.info(f'Available nodes in topology: {topology.get_all_node_names()}')

        # Ищем целевую ноду по имени
        target_node = topology.get_node_by_name(node_name)
        if target_node is None:
            available_names = topology.get_all_node_names()
            logger.error(f'Node "{node_name}" not found. Available nodes: {available_names}')
            return JsonResponse({'error': f'SSH node "{node_name}" not found in topology'}, status=404)

        target_node_id = target_node['id']

        # Включаем все ноды в топологии
        all_node_ids = topology.get_all_node_ids()
        logger.info(f'Starting all nodes in topology: {all_node_ids}')

        # Вспомогательная функция: реавторизация и rejoin при протухании сессии
        def relogin_and_rejoin():
            """Повторная авторизация и присоединение к сессии после 412"""
            nonlocal cookies, xsrf_token
            # Определяем пользователя для повторного логина (мастер или текущий)
            if isinstance(issue, TeamCompetition2TeamsAndUsers) and issue.master_user:
                relogin_user = issue.master_user
            else:
                relogin_user = user
            logger.info(f'Session expired (412), re-logging in as {relogin_user.pnet_login}')
            new_cookies, new_xsrf = pf_login(pnet_url, relogin_user.pnet_login, relogin_user.pnet_password)
            cookies, xsrf_token = new_cookies, new_xsrf
            # Rejoining сессии лабы
            if isinstance(issue, TeamCompetition2TeamsAndUsers) and issue.master_user:
                issue.refresh_from_db()
                if issue.segment_pnet_session_id:
                    from .eveFunctions import join_session
                    join_session(pnet_url, issue.segment_pnet_session_id, cookies, xsrf_token)
                else:
                    create_pnet_lab_session_common(pnet_url, relogin_user.pnet_login, lab_path, cookies, xsrf_token)
            else:
                # Для обычных сессий (TeamCompetition2Team, Competition2User) — rejoin к лабе
                create_pnet_lab_session_common(pnet_url, relogin_user.pnet_login, lab_path, cookies, xsrf_token)

        failed_nodes = []
        for node_id in all_node_ids:
            node_start_success, node_start_message = turn_on_node(pnet_url, node_id, cookies, xsrf_token)
            # При 412 — повторная авторизация и ещё одна попытка
            if not node_start_success and node_start_message == 'SESSION_EXPIRED':
                relogin_and_rejoin()
                node_start_success, node_start_message = turn_on_node(pnet_url, node_id, cookies, xsrf_token)
            # Для целевой ноды при таймауте даём ещё 2 попытки с паузой
            if not node_start_success and node_id == target_node_id and 'timeout' in (node_start_message or '').lower():
                for _ in range(2):
                    time.sleep(5)
                    node_start_success, node_start_message = turn_on_node(pnet_url, node_id, cookies, xsrf_token)
                    if node_start_success:
                        break
            if not node_start_success:
                logger.warning(f'Failed to start node {node_id}: {node_start_message}')
                failed_nodes.append({'node_id': node_id, 'error': node_start_message})
            else:
                logger.info(f'Node {node_id} started successfully')

        # Если не удалось включить целевую ноду, возвращаем ошибку
        if any(failed['node_id'] == target_node_id for failed in failed_nodes):
            logger.error(f'Failed to start target node {target_node_id}')
            return JsonResponse({
                'error': f'Failed to start target node: {next(f["error"] for f in failed_nodes if f["node_id"] == target_node_id)}'
            }, status=500)

        # Для VNC нод — не ждём готовности, Guacamole переподключится сам
        target_node_console = target_node.get('console', '')
        if target_node_console == 'vnc':
            logger.info(f'VNC node {target_node_id}: returning Guacamole URL immediately, it will reconnect when node is ready')

        # Получаем ссылку на Guacamole консоль для целевой ноды
        # cookies/xsrf_token содержат куки мастера для сегментных сессий
        guacamole_url = get_guacamole_url(pnet_url, target_node_id, cookies, xsrf_token)
        if not guacamole_url:
            logger.warning(f'Failed to get guacamole URL, re-logging in and retrying')
            relogin_and_rejoin()
            guacamole_url = get_guacamole_url(pnet_url, target_node_id, cookies, xsrf_token)
        if not guacamole_url:
            return JsonResponse({'error': 'Failed to get console URL'}, status=500)

        response_data = {
            'success': True,
            'node_id': target_node_id,
            'node_name': node_name,
            'guacamole_url': guacamole_url,
            'lab_path': lab_path,
            'all_nodes_started': len(all_node_ids),
            'failed_nodes': failed_nodes if failed_nodes else None
        }
        
        return JsonResponse(response_data)

    except Competition.DoesNotExist:
        return JsonResponse({'error': 'Competition not found'}, status=404)
    except requests.exceptions.Timeout:
        return JsonResponse({'error': 'PNET request timeout'}, status=500)
    except requests.exceptions.ConnectionError:
        return JsonResponse({'error': 'Failed to connect to PNET'}, status=500)
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f'Console session creation error: {str(e)}')
        return JsonResponse({'error': f'Console session creation error: {str(e)}'}, status=500)


@require_POST
def kkz_preview_random(request):
    lab_id = request.GET.get('lab_id')
    num_tasks_str = request.GET.get('num_tasks', '0')
    count_str = request.GET.get('count', '1')

    try:
        num_tasks = int(num_tasks_str)
        count = int(count_str)
    except ValueError:
        return JsonResponse({'error': 'Invalid parameters'}, status=400)

    if num_tasks <= 0:
        return JsonResponse({'sets': [[] for _ in range(count)] if count > 1 else {'tasks': []}})

    try:
        lab = Lab.objects.get(id=lab_id)
    except Lab.DoesNotExist:
        return JsonResponse({'error': 'Lab not found'}, status=400)

    all_tasks = list(lab.options.all())
    if len(all_tasks) < num_tasks:
        return JsonResponse({'error': 'Not enough tasks'}, status=400)

    sets = []
    for _ in range(count):
        sampled = sample_tasks_with_dependencies(all_tasks, num_tasks)
        sets.append([t.id for t in sampled])

    if count == 1:
        return JsonResponse({'tasks': sets[0]})
    else:
        return JsonResponse({'sets': sets})


@require_GET
def kkz_preview_random(request):
    lab_id = request.GET.get("lab_id")
    num_tasks = int(request.GET.get("num_tasks", 0) or 0)
    unified = request.GET.get("unified", "false").lower() in ("1", "true", "yes", "on")
    platoon_ids = request.GET.get("platoon_ids")
    user_ids = request.GET.get("user_ids")
    kkz_id = request.GET.get("kkz_id")
    force_regen = request.GET.get("force_regen", "false").lower() in ("1", "true", "yes", "on")

    selected_task_ids = request.GET.get("selected_tasks")

    if platoon_ids:
        platoon_ids = [int(x) for x in platoon_ids.split(",") if x.strip()]
    else:
        platoon_ids = None

    if user_ids:
        user_ids = [int(x) for x in user_ids.split(",") if x.strip()]
    else:
        user_ids = None

    lab = get_object_or_404(Lab, id=lab_id)
    if selected_task_ids:
        selected_task_ids = [int(x) for x in selected_task_ids.split(",") if x.strip()]
        all_tasks = list(LabTask.objects.filter(lab=lab, id__in=selected_task_ids).order_by('id'))
    else:
        all_tasks = list(LabTask.objects.filter(lab=lab).order_by('id'))

    tasks_json = [{"id": t.id, "description": getattr(t, "description", str(t))} for t in all_tasks]
    task_ids_set = {t.id for t in all_tasks}

    if kkz_id:
        kkz = get_object_or_404(Kkz, id=kkz_id)
        users = list(kkz.get_users())
        unified = kkz.unified_tasks
    else:
        users = set()
        if platoon_ids:
            users.update(User.objects.filter(platoon__in=platoon_ids))
        if user_ids:
            users.update(User.objects.filter(id__in=user_ids))
        users = list(users)

    users = sorted(users, key=lambda u: u.get_full_name() or u.username)
    users_json = [{"id": u.id, "username": u.username, "display": u.get_full_name() or u.username} for u in users]

    assignments = {}
    regenerate = force_regen

    if kkz_id and not force_regen:
        previews = KkzPreview.objects.filter(kkz=kkz, lab=lab)
        previews_dict = {p.user.id: p for p in previews}
        missing_users = [u for u in users if u.id not in previews_dict]
        invalid_previews = False
        for user_id, preview in previews_dict.items():
            preview_task_ids = {t.id for t in preview.tasks.all()}
            if not preview_task_ids.issubset(task_ids_set):
                invalid_previews = True
                break

        if missing_users or invalid_previews or previews.count() != len(users):
            regenerate = True
        else:
            assignments = {str(u.id): [t.id for t in previews_dict[u.id].tasks.all()] for u in users}

    if regenerate or not kkz_id:
        assignments = {}
        if users:
            if unified:
                picked = sample_tasks_with_dependencies(all_tasks, min(num_tasks, len(all_tasks)))
                p_ids = [t.id for t in picked]
                for u in users:
                    assignments[str(u.id)] = p_ids
            else:
                for u in users:
                    sel = sample_tasks_with_dependencies(all_tasks, min(num_tasks, len(all_tasks)))
                    assignments[str(u.id)] = [t.id for t in sel]

    if num_tasks == 1 and all_tasks:
        def _task_has_dependencies(task):
            deps = getattr(task, 'dependencies', None) or ''
            return bool([p for p in deps.split(',') if p.strip()])
        tasks_without_deps = [t for t in all_tasks if not _task_has_dependencies(t)]
        task_by_id = {t.id: t for t in all_tasks}
        for uid in list(assignments.keys()):
            if len(assignments[uid]) == 1:
                tid = assignments[uid][0]
                task = task_by_id.get(tid)
                if task and _task_has_dependencies(task):
                    if tasks_without_deps:
                        assignments[uid] = [random.choice(tasks_without_deps).id]
                    else:
                        assignments[uid] = []

    return JsonResponse({
        "lab": {"id": lab.id, "name": lab.name},
        "tasks": tasks_json,
        "users": users_json,
        "assignments": assignments
    })


@require_POST
def kkz_save_preview(request):
    data = json.loads(request.body)
    kkz_id = data.get('kkz_id')
    lab_id = data.get('lab_id')
    assignments = data.get('assignments', {})

    logger = logging.getLogger(__name__)
    logger.debug(f"kkz_save_preview called: kkz_id={kkz_id}, lab_id={lab_id}, assignments={assignments}")

    kkz = get_object_or_404(Kkz, id=kkz_id)
    lab = get_object_or_404(Lab, id=lab_id)

    if kkz.unified_tasks:
        if assignments:
            first_user_id = next(iter(assignments))
            uniform_task_ids = assignments[first_user_id]
            users = kkz.get_users()
            for user in users:
                preview, created = KkzPreview.objects.update_or_create(
                    kkz=kkz,
                    lab=lab,
                    user=user,
                    defaults={}
                )
                preview.tasks.set(uniform_task_ids)
    else:
        for user_id_str, task_ids in assignments.items():
            try:
                user_id = int(user_id_str)
                user = get_object_or_404(User, id=user_id)
                preview, created = KkzPreview.objects.update_or_create(
                    kkz=kkz,
                    lab=lab,
                    user=user,
                    defaults={}
                )
                preview.tasks.set(task_ids)
            except (ValueError, User.DoesNotExist):
                continue
    return JsonResponse({'status': 'ok'})


def get_labs_for_platoon(request):
    platoon_id = request.GET.get('platoon_id')

    logger = logging.getLogger(__name__)
    logger.debug(f"get_labs_for_platoon called with platoon_id: {platoon_id}")

    if not platoon_id:
        return JsonResponse({'error': 'platoon_id required'}, status=400)

    try:
        platoon = Platoon.objects.get(id=platoon_id)
        logger.debug(f"Found platoon: {platoon.number}, learning_year: {platoon.learning_year}")
    except Platoon.DoesNotExist:
        return JsonResponse({'error': 'Platoon not found'}, status=404)

    all_exam_labs = Lab.objects.filter(lab_type=LabType.EXAM)
    labs = [lab for lab in all_exam_labs if platoon.learning_year in lab.learning_years]
    logger.debug(f"Found {labs} labs for learning_year {platoon.learning_year}")

    labs_data = []
    for lab in labs:
        tasks = LabTask.objects.filter(lab=lab)
        tasks_data = [
            {
                'id': task.id,
                'task_id': (task.task_id or '').strip() or None,
                'description': task.description,
                'dependencies': (task.dependencies or '').strip()
            }
            for task in tasks
        ]

        labs_data.append({
            'id': lab.id,
            'name': lab.name,
            'slug': lab.slug,
            'tasks': tasks_data,
            'default_duration': str(lab.default_duration) if lab.default_duration else None
        })

        logger.debug(f"Lab: {lab.name} with {len(tasks_data)} tasks")

    return JsonResponse({'labs': labs_data})


@require_GET
def get_users_for_platoon(request):
    platoon_ids = request.GET.get("platoon_ids")

    if platoon_ids:
        platoon_ids = [int(x) for x in platoon_ids.split(",") if x.strip()]
    else:
        return JsonResponse({'users': []})

    users = User.objects.filter(platoon__in=platoon_ids).distinct()
    users = sorted(users, key=lambda u: u.get_full_name() or u.username)

    users_json = [{
        "id": u.id,
        "username": u.username,
        "full_name": u.get_full_name() or u.username
    } for u in users]

    return JsonResponse({'users': users_json})


def get_team_or_user_issue(competition, user):
    """
    Получает issue для пользователя в соревновании.
    
    Приоритет: TeamCompetition2Team → TeamCompetition2TeamsAndUsers (сегменты) → Competition2User.
    """
    issue = None
    
    # Сначала пытаемся найти командный issue (если это TeamCompetition)
    if isinstance(competition, TeamCompetition):
        try:
            issue = TeamCompetition2Team.objects.get(competition=competition, team__users=user)
            return issue
        except TeamCompetition2Team.DoesNotExist:
            pass

        issue = TeamCompetition2TeamsAndUsers.objects.filter(
            team_competition=competition
        ).filter(Q(teams__users=user) | Q(users=user)).select_related('master_user').first()
        if issue:
            return issue

    try:
        issue = Competition2User.objects.get(competition=competition, user=user)
    except Competition2User.DoesNotExist:
        pass
    
    return issue

@require_POST
def check_task_answers(request):
    """Проверяет ответы пользователя на задания"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'User not authenticated'}, status=401)
    
    try:
        data = json.loads(request.body)
        competition_slug = data.get('competition_slug')
        answers = data.get('answers', {})  # {task_id: answer_text}
        
        if not competition_slug:
            return JsonResponse({'error': 'Competition slug required'}, status=400)
        
        # Получаем соревнование
        try:
            competition, is_team_competition = get_competition_by_slug(competition_slug)
        except Competition.DoesNotExist:
            return JsonResponse({'error': 'Competition not found'}, status=404)
        
        issue = get_team_or_user_issue(competition, request.user)
        if issue is None:
            return JsonResponse({'error': 'User is not a participant of this competition'}, status=403)
        
        tasks = issue.tasks.all()
        lab = competition.lab
        
        # Проверяем режим проверки заданий
        is_one_attempt = lab.task_checking == TaskChecking.ONE_ATTEMPT
        
        # Получаем список заданий, на которые больше нельзя отвечать
        failed_tasks_list = issue.failed_tasks if issue.failed_tasks else []
        failed_tasks_set = set(failed_tasks_list) if isinstance(failed_tasks_list, list) else set()
        
        # Результаты проверки
        results = {}
        new_failed_tasks = []
        issue_needs_save = False
        
        # Извлекаем флаги из generated_flags
        flags_dict = {}
        if issue.generated_flags:
            if isinstance(issue.generated_flags, dict):
                flags_dict = issue.generated_flags
            elif isinstance(issue.generated_flags, list):
                flags_dict = {
                    item.get('task_id'): item.get('flag')
                    for item in issue.generated_flags
                    if isinstance(item, dict) and item.get('task_id') and item.get('flag')
                }
        
        for task in tasks:
            task_id = str(task.id)
            task_pk = task.id
            
            # Проверяем, есть ли вопрос у задания
            if not task.question or task.question.strip() == '':
                continue
            
            # Проверяем, не находится ли задание в списке failed_tasks
            if task_pk in failed_tasks_set:
                results[task_id] = {
                    'status': 'failed',
                    'message': 'На это задание больше нельзя отвечать'
                }
                continue
            
            # Получаем ответ пользователя
            user_answer = answers.get(task_id, '').strip()
            
            # Если ответ не предоставлен, пропускаем
            if not user_answer:
                results[task_id] = {
                    'status': 'skipped',
                    'message': 'Ответ не предоставлен'
                }
                continue
            
            # Получаем флаг для этого задания (если есть)
            task_flag = flags_dict.get(task.task_id) if task.task_id else None

            is_correct = False
            parsed_choices = None
            try:
                parsed_choices = parse_answer_choices(task.answer) if task.answer else None
            except Exception:
                pass

            if parsed_choices is not None:
                # Задание с выбором ответа: проверяем по индексам
                correct_indices = {opt['index'] for opt in parsed_choices['options'] if opt['correct']}
                user_indices = set()
                for part in user_answer.split(','):
                    part = part.strip()
                    if part.isdigit():
                        user_indices.add(int(part))
                is_correct = user_indices == correct_indices
            else:
                # Свободный ввод: regex/строка и флаг
                correct_answer = task.answer.strip() if task.answer else ''
                if correct_answer:
                    try:
                        is_correct = bool(re.fullmatch(correct_answer, user_answer, re.IGNORECASE))
                    except re.error:
                        is_correct = user_answer.lower() == correct_answer.lower()
                if not is_correct and task_flag:
                    is_correct = user_answer.strip().lower() == task_flag.strip().lower()
            
            results[task_id] = {
                'status': 'correct' if is_correct else 'incorrect',
                'message': 'Верно!' if is_correct else 'Неверно'
            }
            
            # Если ответ правильный, создаем объект Answers
            # Проверяем тип issue, а не is_team_competition, т.к. Competition2User может быть и в TeamCompetition
            answer_filters = {'team': issue.team} if isinstance(issue, TeamCompetition2Team) else {'user': request.user}
            if is_correct:
                # В режиме ONE_ATTEMPT используем транзакцию с блокировкой для предотвращения race condition
                with transaction.atomic():
                    # Блокируем существующий ответ, если он есть
                    Answers.objects.get_or_create(
                        lab=competition.lab,
                        lab_task=task,
                        datetime__lte=competition.finish,
                        datetime__gte=competition.start,
                        defaults={'datetime': timezone.now()},
                        **answer_filters
                    )
                    
            elif is_one_attempt and not is_correct:
                # Если режим ONE_ATTEMPT и ответ неверный, добавляем задание в failed_tasks
                if task_pk not in failed_tasks_set:
                    new_failed_tasks.append(task_pk)
                    failed_tasks_set.add(task_pk)
                    issue_needs_save = True
        
        # Сохраняем обновленный список failed_tasks
        if issue_needs_save:
            issue.failed_tasks = list(failed_tasks_set)
            issue.save(update_fields=['failed_tasks'])
        
        return JsonResponse({
            'success': True,
            'results': results
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Invalid JSON'}, status=400)
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f'Error checking task answers: {str(e)}')
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)


@require_GET
def get_user_tasks_status(request):
    """Возвращает статус выполнения заданий пользователя"""
    if not request.user.is_authenticated:
        return JsonResponse({'error': 'User not authenticated'}, status=401)
    
    competition_slug = request.GET.get('competition_slug')
    
    if not competition_slug:
        return JsonResponse({'error': 'Competition slug required'}, status=400)
    
    try:
        competition, is_team_competition = get_competition_by_slug(competition_slug)
        issue = get_team_or_user_issue(competition, request.user)
        if issue is None:
            return JsonResponse({'error': 'User is not a participant of this competition'}, status=403)
        
        tasks = issue.tasks.all()
        
        # Получаем выполненные задания
        # Проверяем тип issue, а не is_team_competition, т.к. Competition2User может быть и в TeamCompetition
        answer_filters = {'team': issue.team} if isinstance(issue, TeamCompetition2Team) else {'user': request.user} 
        completed_task_ids = set(
            Answers.objects.filter(
                lab=competition.lab,
                lab_task__in=tasks,
                datetime__lte=competition.finish,
                datetime__gte=competition.start,
                **answer_filters
            ).values_list('lab_task_id', flat=True)
        )
        
        # Получаем список заданий, на которые больше нельзя отвечать
        failed_tasks_list = issue.failed_tasks if issue.failed_tasks else []
        failed_tasks_set = set(failed_tasks_list) if isinstance(failed_tasks_list, list) else set()
        
        # Формируем данные о заданиях
        tasks_data = []
        has_questions = False
        
        for idx, task in enumerate(tasks, 1):
            is_completed = task.id in completed_task_ids
            has_question = bool(task.question and task.question.strip())
            is_failed = task.id in failed_tasks_set

            if has_question:
                has_questions = True

            task_payload = {
                'id': task.id,
                'number': idx,
                'description': task.description or '',
                'question': task.question or '',
                'has_question': has_question,
                'done': is_completed,
                'failed': is_failed
            }
            display_choices = task.get_display_choices()
            if display_choices:
                task_payload['answer_type'] = 'single_choice' if display_choices['mode'] == 'single' else 'multiple_choice'
                task_payload['choices'] = [{'id': opt['index'], 'text': opt['text']} for opt in display_choices['options']]
            tasks_data.append(task_payload)
        
        return JsonResponse({
            'success': True,
            'tasks': tasks_data,
            'has_questions': has_questions
        })
        
    except Http404:
        return JsonResponse({'error': 'Competition not found'}, status=404)
    except Exception as e:
        logger = logging.getLogger(__name__)
        logger.error(f'Error getting user tasks status: {str(e)}')
        return JsonResponse({'error': f'Server error: {str(e)}'}, status=500)
